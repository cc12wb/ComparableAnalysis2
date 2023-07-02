import yfinance as yf
import openpyxl
from datetime import datetime
import pandas as pd
import math

def getdatatoXL(companies):
    # create dictionaries for MV, price, div, PE, PB, PS
    mydictMV = {}
    mydictSO = {}
    mydictPreClose = {}
    mydictDiv = {}
    mydictPE = {}
    mydictPB = {}
    mydictPS = {}

    # retrieve MV, price, div, PE, PB, and PS from yfinance
    for i in range(0, len(companies), 1):
        try:
            mydictMV[companies[i]] = yf.Ticker(companies[i]).info['marketCap']
        except:
            print('Could not find market value for ' + companies[i])

    for i in range(0, len(companies), 1):
        try:
            mydictSO[companies[i]] = yf.Ticker(companies[i]).info['sharesOutstanding']
        except:
            print('Could not find shares outstanding for ' + companies[i])

    for i in range(0, len(companies), 1):
        try:
            mydictPreClose[companies[i]] = yf.Ticker(companies[i]).info['previousClose']
        except:
            print('Could not find previous close for ' + companies[i])

    for i in range(0, len(companies), 1):
        try:
            mydictDiv[companies[i]] = yf.Ticker(companies[i]).info['dividendYield']
        except:
            print('Could not find dividend yield for ' + companies[i])

    for i in range(0, len(companies), 1):
        try:
            mydictPE[companies[i]] = yf.Ticker(companies[i]).info['trailingPE']
        except:
            print('Could not find trailing P/E for ' + companies[i])

    for i in range(0, len(companies), 1):
        try:
            mydictPB[companies[i]] = yf.Ticker(companies[i]).info['priceToBook']
        except:
            print('Could not find P/B for ' + companies[i])

    for i in range(0, len(companies), 1):
        try:
            mydictPS[companies[i]] = yf.Ticker(companies[i]).info['priceToSalesTrailing12Months']
        except:
            print('Could not find trailing P/S for ' + companies[i])

    # create excel file databse
    wbCA = openpyxl.Workbook()
    wbCA.create_sheet(index=0, title='Sheet1')
    wsCA = wbCA['Sheet1']

    date = datetime.today().strftime('%Y-%m-%d')
    wbCA.save('Comparable Analysis ' + companies[0] + ' ' + date + '.xlsx')

    # index column
    wsCA.cell(row=2, column=1).value = "Market Cap"
    wsCA.cell(row=3, column=1).value = "Shares Outstanding"
    wsCA.cell(row=4, column=1).value = "Previous Close"
    wsCA.cell(row=5, column=1).value = "Dividend Yield"
    wsCA.cell(row=6, column=1).value = "Trailing PE"
    wsCA.cell(row=7, column=1).value = "P/B"
    wsCA.cell(row=8, column=1).value = "Trailing PS"

    wsCA.cell(row=10, column=1).value = "Income Statement TTM"
    wsCA.cell(row=11, column=1).value = "Revenue"
    wsCA.cell(row=12, column=1).value = "Gross Profit"
    wsCA.cell(row=13, column=1).value = "Operating Income"
    wsCA.cell(row=14, column=1).value = "Interest Expense"
    wsCA.cell(row=15, column=1).value = "Net Income"
    wsCA.cell(row=16, column=1).value = "Interest Coverage Ratio (below 1.5 is questionable)"
    wsCA.cell(row=17, column=1).value = "Gross Profit Margin"
    wsCA.cell(row=18, column=1).value = "Operating Profit Margin"
    wsCA.cell(row=19, column=1).value = "Net Profit Margin"

    wsCA.cell(row=22, column=1).value = "Balance Sheet Statement Latest Quarter"
    wsCA.cell(row=23, column=1).value = "Current Assets"
    wsCA.cell(row=24, column=1).value = "Total Assets"
    wsCA.cell(row=25, column=1).value = "Current Liabilities"
    wsCA.cell(row=26, column=1).value = "Total Liabilities"
    wsCA.cell(row=27, column=1).value = "Total Stockholder Equity"
    wsCA.cell(row=28, column=1).value = "Current Ratio"
    wsCA.cell(row=29, column=1).value = "Debt to Equity Ratio"

    wsCA.cell(row=31, column=1).value = "Cash Flow Statement TTM"
    wsCA.cell(row=32, column=1).value = "Total Cash From Operating Activities"
    wsCA.cell(row=33, column=1).value = "Capital Expenditure"
    wsCA.cell(row=34, column=1).value = "Net Borrowings"
    wsCA.cell(row=35, column=1).value = "Free Cash Flow to Equity"
    wsCA.cell(row=36, column=1).value = "Price to Cash Flow"

    wsCA.cell(row=39, column=1).value = "Implied price from P/E"
    wsCA.cell(row=40, column=1).value = "Implied price from P/B"
    wsCA.cell(row=41, column=1).value = "Implied price from P/S"
    wsCA.cell(row=42, column=1).value = "Implied price from P/CF"
    wsCA.cell(row=43, column=1).value = "Average implied price"

    wsCA.cell(row=46, column=1).value = "Market Cap"
    wsCA.cell(row=47, column=1).value = "Long Term Debt"
    wsCA.cell(row=48, column=1).value = "Short Long Term Debt"
    wsCA.cell(row=49, column=1).value = "Cash"
    wsCA.cell(row=50, column=1).value = "Enterprise Value"
    wsCA.cell(row=51, column=1).value = "Operating Income"
    wsCA.cell(row=52, column=1).value = "Depreciation/Amortization/Depletion"
    wsCA.cell(row=53, column=1).value = "EBITDA"
    wsCA.cell(row=54, column=1).value = "Enterprise Value/EBITDA"

    # write industry in cell A1
    wsCA.cell(row=1, column=1).value = yf.Ticker(companies[0]).info['industry']
    # write MV, price, div, PE, PB, PS to excel database
    for i in range(0, len(companies), 1):
        column = i + 2
        wsCA.cell(row=1, column=column).value = companies[i] + " - " + yf.Ticker(companies[i]).info['shortName']

    for i in range(0, len(companies), 1):
        column = i + 2
        if companies[i] in mydictMV.keys():
            wsCA.cell(row=2, column=column).value = mydictMV[companies[i]]
        if companies[i] in mydictSO.keys():
            wsCA.cell(row=3, column=column).value = mydictSO[companies[i]]
        if companies[i] in mydictPreClose.keys():
            wsCA.cell(row=4, column=column).value = mydictPreClose[companies[i]]
        if companies[i] in mydictDiv.keys():
            wsCA.cell(row=5, column=column).value = mydictDiv[companies[i]]
        if companies[i] in mydictPE.keys():
            wsCA.cell(row=6, column=column).value = mydictPE[companies[i]]
        if companies[i] in mydictPB.keys():
            wsCA.cell(row=7, column=column).value = mydictPB[companies[i]]
        if companies[i] in mydictPS.keys():
            wsCA.cell(row=8, column=column).value = mydictPS[companies[i]]

    # retrieve revenue, GP, OI, and NI from yfinance
    for i in range(0, len(companies), 1):
        column = i + 2
        try:
            companyFinancials = yf.Ticker(companies[i]).quarterly_financials
        except:
            print('Could not find financials for ' + companies[i])

        # find the year of the latest financial statements

        yearList = []
        for col in companyFinancials.columns:
            yearHeader = str(col)
            yearList.append(yearHeader[0:10])
        wsCA.cell(row=10, column=column).value = yearList[0] + " - " + yearList[3]
        # for col in companyFinancials.columns:
        #    yearHeader = str(col)
        #    break
        # wsCA.cell(row=10, column=column).value = int(yearHeader[0:4])

        ttmRevenue = companyFinancials.loc['Total Revenue'][0] + companyFinancials.loc['Total Revenue'][1] + \
                     companyFinancials.loc['Total Revenue'][2] + companyFinancials.loc['Total Revenue'][3]
        wsCA.cell(row=11, column=column).value = ttmRevenue

        ttmGP = companyFinancials.loc['Gross Profit'][0] + companyFinancials.loc['Gross Profit'][1] + \
                         companyFinancials.loc['Gross Profit'][2] + companyFinancials.loc['Gross Profit'][3]
        wsCA.cell(row=12, column=column).value = ttmGP

        ttmOI = companyFinancials.loc['Operating Income'][0] + companyFinancials.loc['Operating Income'][1] + \
                companyFinancials.loc['Operating Income'][2] + companyFinancials.loc['Operating Income'][3]
        wsCA.cell(row=13, column=column).value = ttmOI

        try:
            ttmIE = companyFinancials.loc['Interest Expense'][0] + companyFinancials.loc['Interest Expense'][1] + \
                    companyFinancials.loc['Interest Expense'][2] + companyFinancials.loc['Interest Expense'][3]
            wsCA.cell(row=14, column=column).value = ttmIE
        except:
            wsCA.cell(row=14, column=column).value = None

        ttmNI = companyFinancials.loc['Net Income'][0] + companyFinancials.loc['Net Income'][1] + \
                companyFinancials.loc['Net Income'][2] + companyFinancials.loc['Net Income'][3]
        #print(companyFinancials.loc['Net Income'][0])
        #print(companyFinancials.loc['Net Income'][1])
        #print(companyFinancials.loc['Net Income'][2])
        #print(companyFinancials.loc['Net Income'][3])
        wsCA.cell(row=15, column=column).value = ttmNI

        # iterate over financial statements DataFrame and retrieve Revenue, GP, OI, and NI
        #for x, row in enumerate(companyFinancials.values):
         #   if companyFinancials.index[x] == "Total Revenue":
           #     wsCA.cell(row=11, column=column).value = row[0]
         #   elif companyFinancials.index[x] == "Gross Profit":
           #     wsCA.cell(row=12, column=column).value = row[0]
           # elif companyFinancials.index[x] == "Operating Income":
           #     wsCA.cell(row=13, column=column).value = row[0]
           # elif companyFinancials.index[x] == "Interest Expense":
            #    wsCA.cell(row=14, column=column).value = row[0]
           # elif companyFinancials.index[x] == "Net Income":
             #   wsCA.cell(row=15, column=column).value = row[0]

    # calculate interest coverage, GPM, OPM, and NIM
    for i in range(0, len(companies), 1):
        column = i + 2
        # interest coverage
        if wsCA.cell(row=13, column=column).value != None and wsCA.cell(row=14, column=column).value != None:
            wsCA.cell(row=16, column=column).value = (wsCA.cell(row=15, column=column).value / wsCA.cell(row=14,
                                                                                                         column=column).value) * -1
        # GPM
        if wsCA.cell(row=11, column=column).value != None and wsCA.cell(row=12, column=column).value != None:
            wsCA.cell(row=17, column=column).value = wsCA.cell(row=12, column=column).value / wsCA.cell(row=11,
                                                                                                        column=column).value
        # OPM
        if wsCA.cell(row=13, column=column).value != None and wsCA.cell(row=11, column=column).value != None:
            wsCA.cell(row=18, column=column).value = wsCA.cell(row=13, column=column).value / wsCA.cell(row=11,
                                                                                                        column=column).value
        # NIM
        if wsCA.cell(row=11, column=column).value != None and wsCA.cell(row=15, column=column).value != None:
            wsCA.cell(row=19, column=column).value = wsCA.cell(row=15, column=column).value / wsCA.cell(row=11,
                                                                                                        column=column).value

    # retrieve balance sheet info from yfinance
    for i in range(0, len(companies), 1):
        column = i + 2
        try:
            companyBS = yf.Ticker(companies[i]).quarterly_balancesheet
        except:
            print('Could not find balance sheet for ' + companies[i])

        # find the year of the latest financial statements
        for col in companyBS.columns:
            yearHeaderBS = str(col)
            break
        wsCA.cell(row=22, column=column).value = int(yearHeaderBS[0:4])

        # iterate over balance sheete and retrieve assets, liabilities, and equity
        for x, row in enumerate(companyBS.values):
            if companyBS.index[x] == "Current Assets": #used to be Total Current Assets
                wsCA.cell(row=23, column=column).value = row[0]
            elif companyBS.index[x] == "Total Assets":
                wsCA.cell(row=24, column=column).value = row[0]
            elif companyBS.index[x] == "Current Liabilities": #used to be Total Current Liabilities
                wsCA.cell(row=25, column=column).value = row[0]
            elif companyBS.index[x] == "Long Term Debt": #used to be Total Liab
                wsCA.cell(row=26, column=column).value = row[0]
            elif companyBS.index[x] == "Stockholders Equity": #used to be Total Stockholder Equity
                wsCA.cell(row=27, column=column).value = row[0]

        # calculate current ratio and debt to equity
        for i in range(0, len(companies), 1):
            column = i + 2
            if wsCA.cell(row=23, column=column).value != None and wsCA.cell(row=25, column=column).value != None:
                wsCA.cell(row=28, column=column).value = wsCA.cell(row=23, column=column).value / wsCA.cell(row=25,
                                                                                                            column=column).value
            if wsCA.cell(row=26, column=column).value != None and wsCA.cell(row=27, column=column).value != None:
                wsCA.cell(row=29, column=column).value = wsCA.cell(row=26, column=column).value / wsCA.cell(row=27,
                                                                                                            column=column).value

    # retrieve cash flow info from yfinance
    for i in range(0, len(companies), 1):
        column = i + 2
        try:
            companyCF = yf.Ticker(companies[i]).quarterly_cashflow
        except:
            print('Could not find cash flow statement for ' + companies[i])

        #find the year of the latest financial statements
        yearListCF = []
        for col in companyCF.columns:
            yearHeaderCF = str(col)
            yearListCF.append(yearHeaderCF[0:10])
        wsCA.cell(row=31, column=column).value = yearListCF[0] + " - " + yearListCF[-1]

        try:
            ttmCFO = companyCF.loc['Total Cash From Operating Activities'][0] + \
                     companyCF.loc['Total Cash From Operating Activities'][1] + \
                     companyCF.loc['Total Cash From Operating Activities'][2] + \
                     companyCF.loc['Total Cash From Operating Activities'][3]
            wsCA.cell(row=32, column=column).value = ttmCFO
        except KeyError:
            try:
                ttmCFO = companyCF.loc['Operating Cash Flow'][0] + \
                         companyCF.loc['Operating Cash Flow'][1] + \
                         companyCF.loc['Operating Cash Flow'][2] + \
                         companyCF.loc['Operating Cash Flow'][3]
                wsCA.cell(row=32, column=column).value = ttmCFO
            except:
                wsCA.cell(row=32, column=column).value = None
        try:
            ttmCE = companyCF.loc['Capital Expenditure'][0] + \
                     companyCF.loc['Capital Expenditure'][1] + \
                     companyCF.loc['Capital Expenditure'][2] + \
                     companyCF.loc['Capital Expenditure'][3] #companyCF.loc['Capital Expenditures'][0] + \
            wsCA.cell(row=33, column=column).value = ttmCE
        except KeyError:
            wsCA.cell(row=33, column=column).value = None

        try:
            ttmNB = companyCF.loc['Net Issuance Payments Of Debt'][0] + \
                    companyCF.loc['Net Issuance Payments Of Debt'][1] + \
                    companyCF.loc['Net Issuance Payments Of Debt'][2] + \
                    companyCF.loc['Net Issuance Payments Of Debt'][3]

            #ttmNB = companyCF.loc['Net Borrowings'][0] + \
                    #companyCF.loc['Net Borrowings'][1] + \
                   # companyCF.loc['Net Borrowings'][2] + \
                   # companyCF.loc['Net Borrowings'][3]
            wsCA.cell(row=34, column=column).value = ttmNB
        except:
            wsCA.cell(row=34, column=column).value = None

        # iterate over balance sheete and retrieve assets, liabilities, and equity
        #for x, row in enumerate(companyCF.values):
         #   if companyCF.index[x] == "Total Cash From Operating Activities":
          #      wsCA.cell(row=32, column=column).value = row[0]
          #  elif companyCF.index[x] == "Capital Expenditures":
          #      wsCA.cell(row=33, column=column).value = row[0]
          #  elif companyCF.index[x] == "Net Borrowings":
           #     wsCA.cell(row=34, column=column).value = row[0]

        # calculate FCFE and P/CF
        for i in range(0, len(companies), 1):
            column = i + 2
            if wsCA.cell(row=32, column=column).value == None:
                wsCA.cell(row=32, column=column).value = 0
            if wsCA.cell(row=33, column=column).value == None:
                wsCA.cell(row=33, column=column).value = 0
            if wsCA.cell(row=34, column=column).value == None:
                wsCA.cell(row=34, column=column).value = 0
            #tried to take out above if statement. Don't want to not calculate FCFE just because one item is missing
            wsCA.cell(row=35, column=column).value = wsCA.cell(row=32, column=column).value + wsCA.cell(row=33,
                                                                                                        column=column).value + wsCA.cell(
                                                                                                        row=34, column=column).value
            if wsCA.cell(row=35, column=column).value == 0:
                wsCA.cell(row=35, column=column).value = None
            if wsCA.cell(row=4, column=column).value != None and wsCA.cell(row=32,
                                                                           column=column).value != None and wsCA.cell(
                    row=3, column=column).value != None:
                if wsCA.cell(row=35, column=column).value == None or wsCA.cell(row=35, column=column).value < 0:
                    wsCA.cell(row=36, column=column).value = None
                else:
                    wsCA.cell(row=36, column=column).value = wsCA.cell(row=4, column=column).value / (
                                wsCA.cell(row=35, column=column).value / wsCA.cell(row=3, column=column).value)

    # implied price

    # find the average PE ratio and multiply by the EPS of company
    if wsCA.cell(row=15, column=2).value == None:
        wsCA.cell(row=39, column=2).value = None
    elif wsCA.cell(row=15, column=2).value > 0:
        peRatio = []
        for i in range(1, len(companies), 1):
            column = i + 2
            if wsCA.cell(row=6, column=column).value != None and wsCA.cell(row=6, column=column).value > 0:
                peRatio.append(wsCA.cell(row=6, column=column).value)
        if len(peRatio) < 3:
            wsCA.cell(row=39, column=2).value = None
        else:
            averagePE = Average(peRatio)
            wsCA.cell(row=39, column=2).value = averagePE * (
                        wsCA.cell(row=15, column=2).value / wsCA.cell(row=3, column=2).value)

    # find the average PB ratio and multiply by the book value of company
    pbRatio = []
    for i in range(1, len(companies), 1):
        column = i + 2
        if wsCA.cell(row=7, column=column).value == None:
            continue
        elif wsCA.cell(row=7, column=column).value > 0:
            pbRatio.append(wsCA.cell(row=7, column=column).value)

    if len(pbRatio) < 3:
        wsCA.cell(row=40, column=2).value = None
    else:
        averagePB = Average(pbRatio)
        try:
            wsCA.cell(row=40, column=2).value = averagePB * (
                        wsCA.cell(row=27, column=2).value / wsCA.cell(row=3, column=2).value)
        except:
            wsCA.cell(row=40, column=2).value = None

    # find the average PS ratio and multiply by the sales of company
    psRatio = []
    for i in range(1, len(companies), 1):
        column = i + 2
        if wsCA.cell(row=8, column=column).value == None:
            continue
        elif wsCA.cell(row=8, column=column).value > 0:
            psRatio.append(wsCA.cell(row=8, column=column).value)

    if len(psRatio) < 3:
        wsCA.cell(row=41, column=2).value = None
    else:
        averagePS = Average(psRatio)
        wsCA.cell(row=41, column=2).value = averagePS * (
                    wsCA.cell(row=11, column=2).value / wsCA.cell(row=3, column=2).value)

    # find the average P/CF ratio and multiply by the sales of company
    if wsCA.cell(row=35, column=2).value == None:
        wsCA.cell(row=42, column=2).value = None
    elif wsCA.cell(row=35, column=2).value > 0:
        pcfRatio = []
        for i in range(1, len(companies), 1):
            column = i + 2
            if wsCA.cell(row=36, column=column).value != None and wsCA.cell(row=36, column=column).value > 0:
                pcfRatio.append(wsCA.cell(row=36, column=column).value)

        if len(pcfRatio) < 3:
            wsCA.cell(row=42, column=2).value = None
        else:
            averagePCF = Average(pcfRatio)
            wsCA.cell(row=42, column=2).value = averagePCF * (
                        wsCA.cell(row=35, column=2).value / wsCA.cell(row=3, column=2).value)

    impliedPrice = []
    for i in range(39, 43, 1):
        if wsCA.cell(row=i, column=2).value != None:
            impliedPrice.append(wsCA.cell(row=i, column=2).value)

    if len(impliedPrice) == 0:
        wsCA.cell(row=43, column=2).value = None
    else:
        averageimpliedPrice = Average(impliedPrice)
        wsCA.cell(row=43, column=2).value = averageimpliedPrice


    #calculating EV/EBITDA
    for i in range(0, len(companies), 1):
        column = i + 2
        wsCA.cell(row=46, column=column).value = wsCA.cell(row=2, column=column).value

        try:
            companyBS = yf.Ticker(companies[i]).quarterly_balancesheet
        except (IndexError, KeyError, TypeError):
            print('Could not find balance sheet for ' + companies[i])

        try:
            wsCA.cell(row=47, column=column).value = companyBS.loc['Long Term Debt'][0]
        except (IndexError, KeyError, TypeError):
            wsCA.cell(row=47, column=column).value = 0

        try:
            wsCA.cell(row=48, column=column).value = companyBS.loc['Short Long Term Debt'][0]
        except (IndexError, KeyError, TypeError):
            wsCA.cell(row=48, column=column).value = 0

        try:
            wsCA.cell(row=49, column=column).value = companyBS.loc['Cash And Cash Equivalents'][0] # wsCA.cell(row=49, column=column).value = companyBS.loc['Cash'][0]
        except (IndexError, KeyError, TypeError):
            wsCA.cell(row=49, column=column).value = 0

        if wsCA.cell(row=47, column=column).value == None or math.isnan(wsCA.cell(row=47, column=column).value) or wsCA.cell(row=47, column=column).value == 'nan':
            wsCA.cell(row=47, column=column).value = 0
        if wsCA.cell(row=48, column=column).value == None or math.isnan(wsCA.cell(row=48, column=column).value) or wsCA.cell(row=48, column=column).value == 'nan':
            wsCA.cell(row=48, column=column).value = 0
        if wsCA.cell(row=49, column=column).value == None or math.isnan(wsCA.cell(row=49, column=column).value) or wsCA.cell(row=49, column=column).value == 'nan':
            wsCA.cell(row=49, column=column).value = 0
    # retrieve income statement info from yfinance
    for i in range(0, len(companies), 1):
        column = i + 2
        try:
            companyIS = yf.Ticker(companies[i]).quarterly_financials
        except:
            print('Could not find income statement statement for ' + companies[i])

        try:
            ttmOI = companyIS.loc['Operating Income'][0] + \
                     companyIS.loc['Operating Income'][1] + \
                     companyIS.loc['Operating Income'][2] + \
                     companyIS.loc['Operating Income'][3]
            wsCA.cell(row=51, column=column).value = ttmOI
        except KeyError:
            wsCA.cell(row=51, column=column).value = 0

    # retrieve cash flow info from yfinance
    for i in range(0, len(companies), 1):
        column = i + 2
        try:
            companyCF = yf.Ticker(companies[i]).quarterly_cashflow
        except:
            print('Could not find cash flow statement for ' + companies[i])

        try:
            ttmCE = companyCF.loc['Depreciation Amortization Depletion'][0] + \
                     companyCF.loc['Depreciation Amortization Depletion'][1] + \
                     companyCF.loc['Depreciation Amortization Depletion'][2] + \
                     companyCF.loc['Depreciation Amortization Depletion'][3] #companyCF.loc['Depreciation'][3]
            wsCA.cell(row=52, column=column).value = ttmCE
        except KeyError:
            wsCA.cell(row=52, column=column).value = 0

    #Calculate EV and EBITDA
    for i in range(0, len(companies), 1):
        column = i + 2
        wsCA.cell(row=50, column=column).value = float(wsCA.cell(row=46, column=column).value) + \
                                                 float(wsCA.cell(row=47, column=column).value) + \
                                                 float(wsCA.cell(row=48, column=column).value) - \
                                                 float(wsCA.cell(row=49, column=column).value)

        wsCA.cell(row=53, column=column).value = float(wsCA.cell(row=51, column=column).value) + \
                                                float(wsCA.cell(row=52, column=column).value)

        if wsCA.cell(row=53, column=column).value == 0 or wsCA.cell(row=53, column=column).value == None:
            wsCA.cell(row=54, column=column).value = 0
        else:
            wsCA.cell(row=54, column=column).value = float(wsCA.cell(row=50, column=column).value) / \
                                                float(wsCA.cell(row=53, column=column).value)


    wbCA.save('Comparable Analysis ' + companies[0] + ' ' + date + '.xlsx')




def Average(lst):
    return sum(lst) / len(lst)


def getmarketCap(company):
    try:
        companymarketCap = yf.Ticker(company).info['marketCap']
    except:
        print("Error could not get market cap for " + company)
    return companymarketCap

