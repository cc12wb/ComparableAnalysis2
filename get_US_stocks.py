import yfinance as yf
import openpyxl

#This program finds the industry, sector, and market cap for each ticker in All Major US Exchange Constituents constituents sheet
wb = openpyxl.load_workbook('US_stocks_06252023.xlsx')
ws = wb['Sheet1']

#find last row in index sheet
for i in range(1, 10001, 1):
    if ws.cell(row=i, column=2).value == None:
        lastRow = i
        break

findTicker = None
for x in range(2,lastRow,1):
    try:
        indexTicker = ws.cell(row=x, column=2).value
        findTicker = yf.Ticker(indexTicker)
        ws.cell(row=x, column=3).value = findTicker.info['longName']
        ws.cell(row=x, column=4).value = findTicker.info['industry']
        ws.cell(row=x, column=5).value = findTicker.info['sector']
        ws.cell(row=x,column=6).value = findTicker.info['marketCap']
    except:
        print("Error for ticker " + str(indexTicker))

wb.save('US_stocks_06252023.xlsx')
wb.close()