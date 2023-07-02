import openpyxl
import yfinance as yf
import pandas as pd
from datetime import datetime
import additionalFunctions

if __name__ == "__main__":

    getstock = input('Please select a stock: ')
else:
    exit()

#retrieve stock from yfinance
try:
    getstockinfo = yf.Ticker(getstock)
    stockfinancials = yf.Ticker(getstock).financials
    stockindustry = getstockinfo.info['industry']
    stockmarketcap = getstockinfo.info['marketCap']
    stockinfo = getstockinfo.info

except:
    print('Stock not found')
    exit()

#open index corresponding to the stock selected
if '.TO' in getstock:
    wb = openpyxl.load_workbook('CAD_stocks_06252023.xlsx')
else:
    wb = openpyxl.load_workbook('US_stocks_06252023.xlsx') #wb = openpyxl.load_workbook('Russell 1000 constituents.xlsx')

ws = wb['Sheet1']

#find last row in index sheet
for i in range(1, 10001, 1):
    if ws.cell(row=i, column=2).value == None:
        lastRow = i
        break

#create list of tickers following same industry
companies = []
for i in range(1, lastRow, 1):
    if ws.cell(row=i, column=4).value == stockindustry:
        companies.append(ws.cell(row=i, column=2).value)

#checking list of companies
companies_filtered = []
for i in range(0, len(companies), 1):
    column = i + 2
    try:
        company_is = yf.Ticker(companies[i]).quarterly_financials
        company_bs = yf.Ticker(companies[i]).quarterly_balancesheet
        company_cf = yf.Ticker(companies[i]).quarterly_cashflow
        #if you cant find income statement, remove from list of companies
        if company_is.empty == False and len(company_is.columns) >= 4 and \
                company_bs.empty == False and len(company_bs.columns) >= 4 and \
                company_cf.empty == False and len(company_cf.columns) >= 4:
            companies_filtered.append(companies[i])

    except:
        print('Could not find financials for ' + companies[i])
#count the number of tickers in companies and delete if there are more than 5
del companies_filtered[6:]

#if stock is in companies, delete it and put at the beginning of list
if getstock in companies_filtered:
    indexValue = companies_filtered.index(getstock)
    del companies_filtered[indexValue]
    companies_filtered.insert(0, getstock)
else:
    companies_filtered.insert(0,getstock)

additionalFunctions.getdatatoXL(companies_filtered)








